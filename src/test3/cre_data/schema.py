from __future__ import annotations

import csv
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
import hashlib
import io
import json
from pathlib import Path
import tempfile

import duckdb
from openpyxl import load_workbook

from test3.warehouse.geography import validate_geography
from test3.warehouse.temporal import normalize_period

from .metrics import CORE_PROPERTY_TYPES, get_cre_metric
from .mappings import ImportMappingTemplate, apply_mapping


MAX_CRE_CSV_BYTES = 64 * 1024 * 1024
MAX_CRE_ROWS = 1_000_000
REQUIRED = {"market", "geography_type", "geography_id", "period", "frequency", "property_type", "metric", "value",
            "unit", "source_name", "source_identifier", "source_period", "retrieved_at", "methodology", "vintage",
            "licensing_notes", "verification_status"}
VERIFICATION_STATUSES = {"unverified", "analyst_verified", "rejected"}
SOURCE_CLASSES = {"federal_public", "state_local_public", "academic_open", "brokerage_public_report",
                  "public_brokerage_report", "public_company_filing", "analyst_owned", "user_owned",
                  "licensed_local", "manual_research", "unknown"}
TARGET_CLASSIFICATIONS = {"institutional_target", "market_proxy", "residential_proxy", "context_feature"}


def _hash(value: object) -> str:
    payload = json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
    return hashlib.sha256(payload.encode()).hexdigest()


def _date_or_timestamp(value: object, field: str) -> str:
    text = str(value or "").strip()
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        try:
            parsed = datetime.combine(date.fromisoformat(text), datetime.min.time(), tzinfo=timezone.utc)
        except ValueError as exc:
            raise ValueError(f"{field} must be an ISO date or timestamp") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc).isoformat()


def normalize_cre_record(raw: dict, *, row_number: int) -> dict:
    row = {str(key).strip(): (value.strip() if isinstance(value, str) else value) for key, value in raw.items()}
    missing = sorted(field for field in REQUIRED if row.get(field) in (None, ""))
    if missing:
        raise ValueError(f"missing required fields: {missing}")
    if any(str(value or "").lstrip().startswith(("=", "+", "@")) for value in row.values()):
        raise ValueError("spreadsheet formulas are prohibited")
    property_type = str(row["property_type"]).lower()
    if property_type not in CORE_PROPERTY_TYPES:
        raise ValueError("Phase A supports multifamily, industrial, office, and retail target observations")
    metric = get_cre_metric(str(row["metric"]), property_type)
    unit = str(row["unit"])
    if unit not in metric.units_for(property_type):
        raise ValueError(f"unit {unit!r} is invalid for {property_type}/{metric.metric}")
    methodology = str(row["methodology"])
    if methodology not in metric.methodologies:
        raise ValueError(f"methodology {methodology!r} is invalid for {property_type}/{metric.metric}")
    period = normalize_period(row["period"], str(row["frequency"]))
    source_period = normalize_period(row["source_period"], str(row["frequency"]))
    if source_period.label != period.label:
        raise ValueError("source_period must identify the same period as the observation")
    geography = {name: row.get(name) or None for name in ("geography_type", "geography_id", "state_fips", "county_fips", "cbsa")}
    validate_geography(geography)
    if geography["geography_type"] in {"market", "submarket"} and not str(row.get("market") or "").strip():
        raise ValueError("market name is required for market/submarket observations")
    try:
        value = Decimal(str(row["value"]))
    except InvalidOperation as exc:
        raise ValueError("value must be numeric") from exc
    if not value.is_finite() or (metric.minimum is not None and value < Decimal(str(metric.minimum))) or (metric.maximum is not None and value > Decimal(str(metric.maximum))):
        raise ValueError(f"value is outside the governed range for {metric.metric}")
    verification_status = str(row["verification_status"]).lower()
    if verification_status not in VERIFICATION_STATUSES:
        raise ValueError("verification_status must be unverified, analyst_verified, or rejected")
    source_class = str(row.get("source_class") or "unknown").lower()
    if source_class not in SOURCE_CLASSES:
        raise ValueError("source_class is unsupported")
    target_classification = str(row.get("target_classification") or "institutional_target").lower()
    if target_classification not in TARGET_CLASSIFICATIONS:
        raise ValueError("target_classification is unsupported")
    release_date = None
    if row.get("release_date"):
        release_date = date.fromisoformat(str(row["release_date"])).isoformat()
        if date.fromisoformat(release_date) < period.observation_date:
            raise ValueError("release_date cannot precede the observation period")
    retrieved_at = _date_or_timestamp(row["retrieved_at"], "retrieved_at")
    if release_date and datetime.fromisoformat(retrieved_at).date() < date.fromisoformat(release_date):
        raise ValueError("retrieved_at cannot precede release_date")
    sample_count = None
    if row.get("sample_count") not in (None, ""):
        try:
            sample_count = int(row["sample_count"])
        except (TypeError, ValueError) as exc:
            raise ValueError("sample_count must be a non-negative integer") from exc
        if sample_count < 0:
            raise ValueError("sample_count must be a non-negative integer")
    redistribution = str(row.get("redistribution_permitted") or "unknown").lower()
    if redistribution not in {"yes", "no", "unknown"}:
        raise ValueError("redistribution_permitted must be yes, no, or unknown")
    normalized = {
        "market": str(row["market"]), "geography_type": geography["geography_type"], "geography_id": geography["geography_id"],
        "state_fips": geography["state_fips"], "county_fips": geography["county_fips"], "cbsa": geography["cbsa"],
        "submarket": row.get("submarket") or None, "period": period.label, "observation_date": period.observation_date.isoformat(),
        "frequency": period.period_type, "property_type": property_type, "property_subtype": row.get("property_subtype") or None,
        "metric": metric.metric, "value": format(value, "f"), "unit": unit, "currency": "USD" if unit.startswith("USD") else None,
        "source_name": str(row["source_name"]), "source_identifier": str(row["source_identifier"]),
        "source_period": source_period.label, "release_date": release_date, "retrieved_at": retrieved_at,
        "methodology": methodology, "vintage": str(row["vintage"]), "licensing_notes": str(row["licensing_notes"]),
        "redistribution_permitted": redistribution,
        "verification_status": verification_status, "source_class": source_class, "sample_count": sample_count,
        "target_classification": target_classification,
        "notes": row.get("notes") or None, "source_row": row_number, "raw_row_hash": "sha256:" + _hash(raw),
    }
    normalized["observation_id"] = _hash({key: value for key, value in normalized.items() if key not in {"notes", "source_row"}})
    return normalized


def parse_cre_csv(content: bytes) -> tuple[list[dict], list[dict], dict]:
    if len(content) > MAX_CRE_CSV_BYTES:
        raise ValueError("CRE target file exceeds 64 MiB")
    try:
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    except UnicodeDecodeError as exc:
        raise ValueError("CRE target file must be UTF-8 CSV") from exc
    headers = [str(item).strip() for item in (reader.fieldnames or [])]
    if len(headers) != len(set(headers)):
        raise ValueError("CRE target file contains duplicate headers")
    missing = sorted(REQUIRED - set(headers))
    if missing:
        raise ValueError(f"CRE target file is missing columns: {missing}")
    rows, errors = [], []
    for row_number, raw in enumerate(reader, 2):
        if row_number > MAX_CRE_ROWS + 1:
            raise ValueError("CRE target file exceeds 1,000,000 rows")
        try:
            rows.append(normalize_cre_record(raw, row_number=row_number))
        except ValueError as exc:
            errors.append({"row": row_number, "error": str(exc), "raw_row_hash": "sha256:" + _hash(raw)})
    return rows, errors, {"schema_version": "test3-cre-history/1.0.0", "sha256": hashlib.sha256(content).hexdigest(),
                          "bytes": len(content), "valid_rows": len(rows), "invalid_rows": len(errors), "headers": headers}


def _tabular_rows(content: bytes, suffix: str) -> tuple[list[dict], list[str]]:
    suffix = suffix.lower()
    if len(content) > MAX_CRE_CSV_BYTES:
        raise ValueError("CRE target file exceeds 64 MiB")
    if suffix == ".csv":
        try:
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        except UnicodeDecodeError as exc:
            raise ValueError("CRE target CSV must be UTF-8") from exc
        headers = [str(item).strip() for item in (reader.fieldnames or [])]
        rows = list(reader)
    elif suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=False)
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise ValueError("CRE target workbook is empty") from exc
        headers = [str(cell.value or "").strip() for cell in header_cells]
        rows = []
        for cells in iterator:
            if any(cell.data_type == "f" for cell in cells):
                raise ValueError("spreadsheet formulas are prohibited")
            values = [cell.value for cell in cells]
            if any(value not in (None, "") for value in values):
                rows.append(dict(zip(headers, values, strict=True)))
        workbook.close()
    elif suffix == ".parquet":
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.parquet"; path.write_bytes(content)
            with duckdb.connect(":memory:") as connection:
                result = connection.execute("SELECT * FROM read_parquet(?) LIMIT ?", [str(path), MAX_CRE_ROWS + 1])
                headers = [item[0] for item in result.description]
                rows = [dict(zip(headers, values, strict=True)) for values in result.fetchall()]
    else:
        raise ValueError("CRE target imports support CSV, XLSX, or Parquet")
    if not headers or len(headers) != len(set(headers)) or any(not header for header in headers):
        raise ValueError("CRE target file contains empty or duplicate headers")
    if len(rows) > MAX_CRE_ROWS:
        raise ValueError("CRE target file exceeds 1,000,000 rows")
    return rows, headers


def parse_cre_file(content: bytes, *, suffix: str,
                   mapping: ImportMappingTemplate | None = None) -> tuple[list[dict], list[dict], dict]:
    """Parse canonical or mapped CSV/XLSX/Parquet without trusting source-file assertions."""
    source_rows, headers = _tabular_rows(content, suffix)
    mapped = apply_mapping(source_rows, mapping) if mapping else source_rows
    missing = sorted(REQUIRED - set(mapped[0] if mapped else ()))
    if missing:
        raise ValueError(f"CRE target file is missing columns: {missing}")
    rows, errors = [], []
    for row_number, raw in enumerate(mapped, 2):
        try:
            rows.append(normalize_cre_record(raw, row_number=row_number))
        except ValueError as exc:
            errors.append({"row": row_number, "error": str(exc), "raw_row_hash": "sha256:" + _hash(raw)})
    return rows, errors, {
        "schema_version": "test3-cre-history/1.1.0", "sha256": hashlib.sha256(content).hexdigest(),
        "bytes": len(content), "valid_rows": len(rows), "invalid_rows": len(errors), "headers": headers,
        "format": suffix.lower().lstrip("."),
        "mapping_template": ({"template_id": mapping.template_id, "version": mapping.version} if mapping else None),
    }
