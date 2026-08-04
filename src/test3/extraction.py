from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass, replace
from pathlib import Path

import pypdfium2 as pdfium
from openpyxl import load_workbook

from .normalization import number
from .local_ocr import available as ocr_available, extract as ocr_extract, validate_image

MAX_XLSX_ENTRIES = 250
MAX_XLSX_XML_BYTES = 25 * 1024 * 1024
MAX_XLSX_TOTAL_BYTES = 100 * 1024 * 1024
MAX_XLSX_ROWS = 100_000
MAX_XLSX_CELLS = 2_000_000


@dataclass(frozen=True)
class Candidate:
    field: str
    raw: str
    normalized: str | None
    page: int | None
    excerpt: str
    confidence: float
    method: str
    bbox: tuple[float, float, float, float] | None = None


FIELD_PATTERNS = {
    "property_name": r"(?im)^property(?: name)?\s*[:,-]\s*(.+)$",
    "address": r"(?im)^address\s*[:,-]\s*(.+)$",
    "asking_price": r"(?i)(?:asking|purchase) price\s*[:$ ]+([($0-9,.]+)",
    "broker_stated_noi": r"(?i)(?:broker[- ]stated )?noi\s*[:$ ]+([($0-9,.]+)",
    "broker_stated_cap_rate": r"(?i)(?:cap(?:italization)? rate)\s*[: ]+([0-9.]+%?)",
    "rentable_square_feet": r"(?i)(?:rentable (?:area|square feet)|rsf)\s*[: ]+([0-9,.]+)",
    "occupancy": r"(?i)occupancy\s*[: ]+([0-9.]+%?)",
    "unit_count": r"(?i)(?:unit count|units)\s*[: ]+([0-9,]+)",
    "loan_amount": r"(?i)loan amount\s*[:$ ]+([($0-9,.]+)",
    "interest_rate": r"(?i)(?:all-in )?interest rate\s*[: ]+([0-9.]+%?)",
}


def extract_text_candidates(text: str, page: int = 1, method: str = "deterministic_regex_v1") -> list[Candidate]:
    candidates = []
    for field, pattern in FIELD_PATTERNS.items():
        match = re.search(pattern, text)
        if not match:
            continue
        raw = match.group(1).strip()[:500]
        numeric = number(raw)
        normalized = str(numeric) if numeric is not None else raw
        start, end = max(0, match.start() - 60), min(len(text), match.end() + 60)
        candidates.append(Candidate(field, raw, normalized, page, text[start:end].strip(), 0.82, method))
    return candidates


def parse_csv(content: bytes) -> tuple[list[list[str]], list[Candidate]]:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    candidates = []
    if rows:
        headers = [re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_") for value in rows[0]]
        for row_index, row in enumerate(rows[1:], 2):
            for index, raw in enumerate(row):
                if raw.strip() and index < len(headers):
                    candidates.append(Candidate(f"row.{row_index}.{headers[index] or index}", raw, str(number(raw)) if number(raw) is not None else raw.strip(), 1, f"CSV row {row_index}, column {index + 1}", 0.95, "csv_cell_v1", (index, row_index - 1, index + 1, row_index)))
    return rows, candidates


def parse_xlsx(content: bytes) -> tuple[list[list[str]], list[Candidate]]:
    # XLSX is a ZIP of XML. Values only are read; formulas, macros, links and scripts are never evaluated.
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ENTRIES:
            raise ValueError("XLSX archive contains too many entries")
        if sum(entry.file_size for entry in entries) > MAX_XLSX_TOTAL_BYTES:
            raise ValueError("XLSX expanded content exceeds the safety limit")
        for entry in entries:
            if entry.file_size > MAX_XLSX_XML_BYTES:
                raise ValueError("XLSX entry exceeds the safety limit")
            if entry.compress_size and entry.file_size / entry.compress_size > 200:
                raise ValueError("XLSX compression ratio is unsafe")
        names = set(archive.namelist())
        sheet_name = next((name for name in sorted(names) if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")), None)
        if not sheet_name:
            return [], []
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
    try:
        sheet = workbook.worksheets[0]
        if sheet.max_row > MAX_XLSX_ROWS or sheet.max_column * sheet.max_row > MAX_XLSX_CELLS:
            raise ValueError("XLSX dimensions exceed the safety limit")
        rows = [["" if cell.value is None else str(cell.value) for cell in row] for row in sheet.iter_rows()]
    finally:
        workbook.close()
    candidates = []
    if rows:
        headers = [re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_") for value in rows[0]]
        for row_index, row in enumerate(rows[1:], 2):
            for index, raw in enumerate(row):
                if not raw.strip() or index >= len(headers):
                    continue
                formula = raw.startswith("=")
                candidates.append(Candidate(f"row.{row_index}.{headers[index] or index}", raw, None if formula else (str(number(raw)) if number(raw) is not None else raw.strip()), 1, f"XLSX sheet {sheet.title!r}, cell row {row_index}, column {index + 1}", 0.4 if formula else 0.95, "xlsx_formula_not_evaluated_v2" if formula else "xlsx_cell_v2", (index, row_index - 1, index + 1, row_index)))
    return rows, candidates


def extract_selectable_pdf_text(content: bytes) -> str:
    # Conservative fallback for simple PDFs. Complex PDFs remain visibly flagged for local OCR/manual review.
    chunks = []
    for match in re.finditer(rb"\(([^()]*)\)\s*Tj", content, re.DOTALL):
        raw = re.sub(rb"\\([()\\])", rb"\1", match.group(1))
        chunks.append(raw.decode("latin-1", errors="replace"))
    return "\n".join(chunks)


def extract_pdf_candidates(content: bytes) -> tuple[list[Candidate], int]:
    document = pdfium.PdfDocument(content)
    candidates, text_pages = [], 0
    try:
        for page_index in range(len(document)):
            page = document[page_index]
            text_page = page.get_textpage()
            try:
                text = text_page.get_text_range()
                if text.strip():
                    text_pages += 1
                for candidate in extract_text_candidates(text, page_index + 1, "pdfium_text_v2"):
                    bbox = None
                    searcher = text_page.search(candidate.raw)
                    try:
                        match = searcher.get_next()
                        if match:
                            start, count = match
                            boxes = [text_page.get_charbox(index) for index in range(start, start + count)]
                            left, bottom, right, top = min(box[0] for box in boxes), min(box[1] for box in boxes), max(box[2] for box in boxes), max(box[3] for box in boxes)
                            width, height = page.get_size()
                            bbox = (left / width, (height - top) / height, (right - left) / width, (top - bottom) / height)
                    finally:
                        searcher.close()
                    candidates.append(replace(candidate, bbox=bbox))
            finally:
                text_page.close()
                page.close()
    finally:
        document.close()
    return candidates, text_pages


def _ocr_candidates(image_content: bytes, suffix: str, page_number: int) -> list[Candidate]:
    width, height, _ = validate_image(image_content)
    result = ocr_extract(image_content, suffix)
    output = []
    for candidate in extract_text_candidates(result.text, page_number, result.engine):
        wanted = [re.sub(r"\W", "", token).lower() for token in candidate.raw.split()]
        words = result.words
        matched = []
        for start in range(len(words)):
            actual = [re.sub(r"\W", "", item["text"]).lower() for item in words[start:start + len(wanted)]]
            if actual == wanted:
                matched = words[start:start + len(wanted)]
                break
        bbox = None
        if matched:
            boxes = [item["bbox"] for item in matched]
            left, top, right, bottom = min(box[0] for box in boxes), min(box[1] for box in boxes), max(box[2] for box in boxes), max(box[3] for box in boxes)
            bbox = (left / width, top / height, (right - left) / width, (bottom - top) / height)
        confidence = min(candidate.confidence, sum(item["confidence"] for item in matched) / len(matched)) if matched else min(candidate.confidence, 0.65)
        output.append(replace(candidate, bbox=bbox, confidence=confidence))
    return output


def ocr_pdf_candidates(content: bytes) -> list[Candidate]:
    document = pdfium.PdfDocument(content)
    candidates = []
    try:
        if len(document) > 250:
            raise ValueError("PDF page count exceeds the OCR safety limit")
        for page_index in range(len(document)):
            page = document[page_index]
            try:
                width, height = page.get_size()
                scale = min(2.0, 10_000 / max(width, height))
                if width * height * scale * scale > 80_000_000:
                    raise ValueError("Rendered OCR page exceeds the pixel safety limit")
                bitmap = page.render(scale=scale)
                try:
                    image = bitmap.to_pil(); stream = io.BytesIO(); image.save(stream, format="PNG")
                finally:
                    bitmap.close()
                candidates.extend(_ocr_candidates(stream.getvalue(), ".png", page_index + 1))
            finally:
                page.close()
    finally:
        document.close()
    return candidates


def process(filename: str, mime: str, content: bytes) -> tuple[str, list[Candidate], str | None]:
    if mime == "text/csv":
        _, candidates = parse_csv(content)
        return "extracted", candidates, None
    if mime.endswith("spreadsheetml.sheet"):
        try:
            _, candidates = parse_xlsx(content)
            return "extracted", candidates, None
        except Exception as error:
            return "failed", [], f"Spreadsheet could not be parsed: {error}"
    if mime == "application/pdf":
        try:
            candidates, text_pages = extract_pdf_candidates(content)
        except Exception as error:
            text = extract_selectable_pdf_text(content)
            if text.strip():
                return "needs_review", extract_text_candidates(text), f"Mature PDF parser rejected the file; conservative fallback used: {type(error).__name__}"
            return "failed", [], f"PDF could not be safely parsed: {type(error).__name__}"
        if not text_pages:
            if not ocr_available():
                return "needs_review", [], "No selectable text found. Install local Tesseract to enable scanned-page OCR."
            try:
                candidates = ocr_pdf_candidates(content)
                return ("extracted" if candidates else "needs_review"), candidates, None if candidates else "Local OCR completed but no supported fields were identified."
            except (RuntimeError, ValueError, OSError) as error:
                return "failed", [], str(error)
        return "extracted" if candidates else "needs_review", candidates, None if candidates else "Text was extracted but no supported fields were identified."
    if mime.startswith("image/"):
        try:
            validate_image(content)
        except ValueError as error:
            return "failed", [], str(error)
        if not ocr_available():
            return "needs_review", [], "Image verified; optional local Tesseract OCR is not installed."
        try:
            candidates = _ocr_candidates(content, Path(filename).suffix.lower(), 1)
            return ("extracted" if candidates else "needs_review"), candidates, None if candidates else "OCR completed but no supported fields were identified."
        except (RuntimeError, ValueError, OSError) as error:
            return "failed", [], str(error)
    return "failed", [], "Unsupported processor"

