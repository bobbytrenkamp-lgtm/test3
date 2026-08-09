from __future__ import annotations

import csv
import json
from decimal import Decimal
import re

from openpyxl import load_workbook

from .base import PublicDataRequest, PublicDataSource, RawSnapshot, canonical_row


class BLSLAUS(PublicDataSource):
    source_id, dataset_id, domain = "bls_laus_ces", "laus_county", "labor"
    normalizer_version = "2.0.0"
    allowed_hosts = ("download.bls.gov", "api.bls.gov", "www.bls.gov", "data.bls.gov")
    SERIES_SUFFIXES = {"03": ("unemployment_rate", "percent"), "04": ("unemployment", "persons"),
                       "05": ("employment", "persons"), "06": ("labor_force", "persons")}
    CHUNKS = ("00-04", "05-09", "10-14", "15-19", "20-24", "25-29", "90-94", "95-99")
    STATE_FILES = {
        "01":"7.Alabama","02":"8.Alaska","04":"9.Arizona","05":"10.Arkansas","06":"11.California","08":"12.Colorado",
        "09":"13.Connecticut","10":"14.Delaware","11":"15.DC","12":"16.Florida","13":"17.Georgia","15":"18.Hawaii",
        "16":"19.Idaho","17":"20.Illinois","18":"21.Indiana","19":"22.Iowa","20":"23.Kansas","21":"24.Kentucky",
        "22":"25.Louisiana","23":"26.Maine","24":"27.Maryland","25":"28.Massachusetts","26":"29.Michigan",
        "27":"30.Minnesota","28":"31.Mississippi","29":"32.Missouri","30":"33.Montana","31":"34.Nebraska",
        "32":"35.Nevada","33":"36.NewHampshire","34":"37.NewJersey","35":"38.NewMexico","36":"39.NewYork",
        "37":"40.NorthCarolina","38":"41.NorthDakota","39":"42.Ohio","40":"43.Oklahoma","41":"44.Oregon",
        "42":"45.Pennsylvania","72":"46.PuertoRico","44":"47.RhodeIsland","45":"48.SouthCarolina",
        "46":"49.SouthDakota","47":"50.Tennessee","48":"51.Texas","49":"52.Utah","50":"53.Vermont",
        "51":"54.Virginia","53":"56.Washington","54":"57.WestVirginia","55":"58.Wisconsin","56":"59.Wyoming",
    }
    NATIONAL_SERIES = {"LNS12000000": ("employment", "thousands_persons"),
                       "LNS11000000": ("labor_force", "thousands_persons"),
                       "LNS13000000": ("unemployment", "thousands_persons"),
                       "LNS14000000": ("unemployment_rate", "percent")}
    ANNUAL_FIELDS = {
        "laborforce": ("labor_force", "persons", "06"),
        "employed": ("employment", "persons", "05"),
        "unemployed": ("unemployment", "persons", "04"),
        "unemploymentrate": ("unemployment_rate", "percent", "03"),
    }
    QCEW_FIELDS = {
        "annual_avg_emplvl": ("covered_employment", "persons"),
        "annual_avg_estabs_count": ("annual_average_establishments", "establishments"),
        "total_annual_wages": ("total_annual_wages", "USD_current"),
        "annual_avg_wkly_wage": ("average_weekly_wage", "USD_current_per_week"),
        "avg_annual_pay": ("average_annual_pay", "USD_current_per_year"),
    }

    def discover(self, request: PublicDataRequest) -> list[str]:
        if request.parameters.get("local_file"):
            source_url = request.parameters.get("source_url")
            if not source_url:
                raise ValueError("BLS local evidence requires --source-url")
            return [str(source_url)]
        if request.parameters.get("annual_county"):
            year = request.to_year or request.from_year
            if year is None or year < 1990 or year > 2099:
                raise ValueError("BLS annual county retrieval requires one supported year")
            return [f"https://www.bls.gov/lau/laucnty{year % 100:02d}.xlsx"]
        if request.parameters.get("qcew_year"):
            year = int(request.parameters["qcew_year"])
            if year < 1975 or year > 2099:
                raise ValueError("BLS QCEW annual retrieval requires a supported year")
            return [f"https://data.bls.gov/cew/data/api/{year}/a/industry/10.csv"]
        series = request.parameters.get("series")
        if series:
            if series not in self.NATIONAL_SERIES: raise ValueError("unsupported governed BLS national series")
            return [f"https://api.bls.gov/publicAPI/v1/timeseries/data/{series}"]
        state = request.parameters.get("state")
        if state:
            if state not in self.STATE_FILES: raise ValueError("unsupported BLS state FIPS")
            return [f"https://download.bls.gov/pub/time.series/la/la.data.{self.STATE_FILES[state]}"]
        chunk = request.parameters.get("chunk", "00-04")
        if chunk not in self.CHUNKS:
            raise ValueError("unsupported BLS bounded download chunk")
        return [f"https://download.bls.gov/pub/time.series/la/la.data.0.CurrentU{chunk}"]

    def normalize(self, snapshot: RawSnapshot):
        if snapshot.content_path.suffix.lower() == ".xlsx":
            yield from self._normalize_annual_county(snapshot)
            return
        if snapshot.request_parameters.get("parameters", {}).get("qcew_year"):
            yield from self._normalize_qcew(snapshot)
            return
        if "api.bls.gov" in snapshot.final_url:
            payload = json.loads(snapshot.content_path.read_text(encoding="utf-8-sig"))
            series_values = payload.get("Results", {}).get("series", []) if isinstance(payload, dict) else []
            if payload.get("status") != "REQUEST_SUCCEEDED" or len(series_values) != 1:
                raise ValueError("BLS public API response failed or changed schema")
            series_id = series_values[0].get("seriesID")
            if series_id not in self.NATIONAL_SERIES: raise ValueError("BLS returned an unexpected series")
            metric, unit = self.NATIONAL_SERIES[series_id]
            for row_number, raw in enumerate(series_values[0].get("data", []), 1):
                if raw.get("period") == "M13": continue
                month = int(str(raw["period"])[1:]); year = int(raw["year"])
                value = Decimal(str(raw["value"])) * 1000 if unit == "thousands_persons" else Decimal(str(raw["value"]))
                canonical_unit = "persons" if unit == "thousands_persons" else unit
                yield canonical_row(snapshot, series=series_id, geography_type="national", geography_id="US",
                                    observation_date=f"{year:04d}-{month:02d}", period_type="monthly", metric=metric,
                                    value=value, unit=canonical_unit, source_row=row_number, raw=raw,
                                    methodology="BLS CPS national series; published thousands converted to persons by deterministic x1000; seasonal adjustment retained by series ID.")
            return
        with snapshot.content_path.open(encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream, delimiter="\t")
            required = {"series_id", "year", "period", "value"}
            if not reader.fieldnames or not required.issubset({x.strip() for x in reader.fieldnames}):
                raise ValueError("BLS LAUS schema changed")
            for row_number, source in enumerate(reader, 2):
                raw = {str(key).strip(): str(value).strip() for key, value in source.items() if key is not None}
                if raw["period"] == "M13":
                    continue
                year = int(raw["year"])
                start, end = snapshot.request_parameters.get("from_year"), snapshot.request_parameters.get("to_year")
                if start and year < start or end and year > end:
                    continue
                series_id = raw["series_id"]
                match = self.SERIES_SUFFIXES.get(series_id[-2:])
                if not match or len(series_id) < 18 or not series_id.startswith("LAUCN"):
                    continue
                fips = series_id[5:10]
                month = int(raw["period"][1:])
                metric, unit = match
                yield canonical_row(snapshot, series=series_id, geography_type="county", geography_id=fips,
                                    observation_date=f"{year:04d}-{month:02d}", period_type="monthly", metric=metric,
                                    value=raw["value"], unit=unit, source_row=row_number, state_fips=fips[:2],
                                    county_fips=fips, raw=raw, methodology="BLS LAUS county series; preliminary/revised status remains in raw evidence.")

    @staticmethod
    def _header(value: object) -> str:
        return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

    def _normalize_annual_county(self, snapshot: RawSnapshot):
        workbook = load_workbook(snapshot.content_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header = None
            header_row = None
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                normalized = [self._header(item) for item in values]
                if {"statefipscode", "countyfipscode", "year", "laborforce", "employed", "unemployed"}.issubset(set(normalized)):
                    header, header_row = normalized, row_number
                    break
            if header is None:
                raise ValueError("BLS annual county workbook schema changed")
            indexes = {name: header.index(name) for name in set(header) if name}
            rate_name = next((name for name in indexes if name.startswith("unemploymentrate")), None)
            if rate_name is None:
                raise ValueError("BLS annual county workbook is missing unemployment rate")
            field_names = {**self.ANNUAL_FIELDS, rate_name: self.ANNUAL_FIELDS["unemploymentrate"]}
            for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                if not any(value is not None for value in values):
                    continue
                def cell(name):
                    index = indexes.get(name)
                    return values[index] if index is not None and index < len(values) else None
                state = str(cell("statefipscode") or "").split(".", 1)[0].zfill(2)
                county = str(cell("countyfipscode") or "").split(".", 1)[0].zfill(3)
                year_value = cell("year")
                if not (state.isdigit() and county.isdigit() and len(state) == 2 and len(county) == 3 and year_value is not None):
                    continue
                year = int(year_value)
                start, end = snapshot.request_parameters.get("from_year"), snapshot.request_parameters.get("to_year")
                if start and year < start or end and year > end:
                    continue
                fips = state + county
                for source_field, (metric, unit, suffix) in field_names.items():
                    if source_field == "unemploymentrate" and rate_name != source_field:
                        continue
                    value = cell(source_field)
                    if value in (None, "", "N.A."):
                        continue
                    cleaned = str(value).replace(",", "").replace("%", "").strip()
                    raw = {"row": row_number, "field": source_field, "value": value,
                           "state_fips": state, "county_fips": county, "year": year}
                    yield canonical_row(
                        snapshot, series=f"LAUCN{fips}0000000{suffix}", geography_type="county", geography_id=fips,
                        observation_date=str(year), period_type="annual", metric=metric, value=cleaned, unit=unit,
                        source_row=row_number, state_fips=state, county_fips=fips, raw=raw,
                        methodology="BLS LAUS county annual average workbook; annual values remain annual and are not expanded to months.",
                    )
        finally:
            workbook.close()

    def _normalize_qcew(self, snapshot: RawSnapshot):
        with snapshot.content_path.open(encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            required = {"area_fips", "own_code", "industry_code", "year", "qtr", "disclosure_code", *self.QCEW_FIELDS}
            if not reader.fieldnames or not required.issubset(reader.fieldnames):
                raise ValueError("BLS QCEW annual CSV schema changed")
            found = 0
            for row_number, raw in enumerate(reader, 2):
                area = str(raw.get("area_fips") or "").strip()
                if (not re.fullmatch(r"\d{5}", area) or area.endswith("000") or
                        str(raw.get("own_code")).strip() != "0" or str(raw.get("industry_code")).strip() != "10" or
                        str(raw.get("qtr")).strip().upper() != "A" or str(raw.get("disclosure_code") or "").strip().upper() == "N"):
                    continue
                year = int(raw["year"])
                for field, (metric, unit) in self.QCEW_FIELDS.items():
                    value = str(raw.get(field) or "").replace(",", "").strip()
                    if value in ("", "-"):
                        continue
                    found += 1
                    yield canonical_row(
                        snapshot, series=f"QCEW:{area}:10:0:{field}", geography_type="county", geography_id=area,
                        observation_date=str(year), period_type="annual", metric=metric, value=value, unit=unit,
                        source_row=row_number, state_fips=area[:2], county_fips=area,
                        raw={"field": field, **raw}, methodology=("BLS Quarterly Census of Employment and Wages annual average, "
                            "all covered industries and ownerships; disclosure-suppressed records are omitted."),
                    )
            if found == 0:
                raise ValueError("BLS QCEW response contains no governed county total rows")
