from __future__ import annotations

import re

from openpyxl import load_workbook

from .base import PublicDataRequest, PublicDataSource, RawSnapshot, canonical_row


class CensusCBSACrosswalk(PublicDataSource):
    """Official OMB/Census county components of CBSAs, retained by delineation vintage."""

    source_id, dataset_id, domain = "census_cbsa_crosswalk", "county_cbsa_2023", "geography"
    allowed_hosts = ("www2.census.gov", "www.census.gov")
    VINTAGES = {
        "2023": (
            "2023-07-21",
            "https://www2.census.gov/programs-surveys/metro-micro/geographies/reference-files/2023/"
            "delineation-files/list1_2023.xlsx",
        ),
    }

    def discover(self, request: PublicDataRequest) -> list[str]:
        vintage = str(request.parameters.get("vintage", "2023"))
        if vintage not in self.VINTAGES:
            raise ValueError("unsupported governed Census CBSA delineation vintage")
        return [self.VINTAGES[vintage][1]]

    @staticmethod
    def _header(value: object) -> str:
        return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

    def normalize(self, snapshot: RawSnapshot):
        vintage = str(snapshot.request_parameters.get("parameters", {}).get("vintage", "2023"))
        effective_date, _ = self.VINTAGES[vintage]
        workbook = load_workbook(snapshot.content_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header = None
            header_row = None
            required = {"cbsacode", "cbsatitle", "countycountyequivalent", "statename", "fipsstatecode", "fipscountycode"}
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                normalized = [self._header(value) for value in values]
                if required.issubset(set(normalized)):
                    header, header_row = normalized, row_number
                    break
            if header is None:
                raise ValueError("Census CBSA delineation workbook schema changed")
            indexes = {name: header.index(name) for name in required}
            for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                def cell(name):
                    index = indexes[name]
                    return values[index] if index < len(values) else None
                cbsa = str(cell("cbsacode") or "").split(".", 1)[0].zfill(5)
                state = str(cell("fipsstatecode") or "").split(".", 1)[0].zfill(2)
                county = str(cell("fipscountycode") or "").split(".", 1)[0].zfill(3)
                if not (cbsa.isdigit() and state.isdigit() and county.isdigit()
                        and len(cbsa) == 5 and len(state) == 2 and len(county) == 3):
                    continue
                county_fips = state + county
                raw = {
                    "cbsa": cbsa, "cbsa_name": cell("cbsatitle"), "county_fips": county_fips,
                    "county_name": cell("countycountyequivalent"), "state_name": cell("statename"),
                    "effective_date": effective_date, "vintage": vintage,
                }
                yield canonical_row(
                    snapshot, series=f"OMB_BULLETIN_23_01_{cbsa}", geography_type="county",
                    geography_id=county_fips, observation_date=effective_date, period_type="irregular",
                    metric="county_cbsa_membership", value="1", unit="membership", source_row=row_number,
                    state_fips=state, county_fips=county_fips, cbsa=cbsa, raw=raw,
                    methodology=(f"County component of CBSA {cbsa} under the July 2023 OMB delineation; "
                                 f"effective {effective_date}. Membership is categorical, vintage-specific, and not additive."),
                )
        finally:
            workbook.close()
