from __future__ import annotations

import re

from openpyxl import load_workbook

from .base import PublicDataRequest, PublicDataSource, RawSnapshot, canonical_row


SERIES = {
    "rental_vacancy_rate": (
        "https://www.census.gov/housing/hvs/data/histtab1.xlsx",
        "rental_vacancy_rate", "percent", 1956,
        "CPS/HVS quarterly rental vacancy rate for all U.S. rental housing; this is not an institutional multifamily market vacancy rate.",
    ),
    "median_asking_rent_vacant_units": (
        "https://www.census.gov/housing/hvs/data/histtab11.xlsx",
        "median_asking_rent_vacant_units", "USD_current_per_month", 1988,
        "CPS/HVS median asking rent for vacant U.S. rental units offered for rent; this is not an institutional effective-rent or brokerage asking-rent series.",
    ),
}


class CensusHousingVacancySurvey(PublicDataSource):
    source_id, dataset_id, domain = "census_hvs", "housing_vacancy_survey", "housing"
    normalizer_version = "1.0.0"
    allowed_hosts = ("www.census.gov", "census.gov")

    def discover(self, request: PublicDataRequest) -> list[str]:
        series = request.parameters.get("series", "rental_vacancy_rate")
        if series not in SERIES:
            raise ValueError(f"unsupported governed HVS series: {series}")
        return [SERIES[series][0]]

    def normalize(self, snapshot: RawSnapshot):
        requested = snapshot.request_parameters["parameters"].get("series", "rental_vacancy_rate")
        _, metric, unit, first_year, methodology = SERIES[requested]
        start = snapshot.request_parameters.get("from_year") or first_year
        end = snapshot.request_parameters.get("to_year") or 2100
        workbook = load_workbook(snapshot.content_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            year = None
            vintage = "published"
            for row_number, raw in enumerate(sheet.iter_rows(values_only=True), 1):
                cells = tuple(raw)
                if requested == "median_asking_rent_vacant_units" and any(str(value).strip().startswith("Table 11B") for value in cells if value is not None):
                    break
                label = str(cells[0] or "").strip()
                year_match = re.match(r"^(\d{4})(.*)$", label)
                if year_match:
                    year = int(year_match.group(1))
                    vintage = "revised" if "r" in year_match.group(2).lower() else "published"
                    continue
                quarter_match = re.match(r"^([1-4])(?:st|nd|rd|th)", label.lower())
                if year is None or not quarter_match or not start <= year <= end or len(cells) < 2:
                    continue
                value = cells[1]
                if value is None or str(value).strip().upper() in {"", "NA", "N/A", "N.A."}:
                    continue
                quarter = int(quarter_match.group(1))
                evidence = {"year": year, "quarter": quarter, "value": value, "vintage_label": vintage}
                row = canonical_row(
                    snapshot, series=f"CPS_HVS_{requested}_US_{vintage}", geography_type="national",
                    geography_id="US", observation_date=f"{year}-Q{quarter}", period_type="quarterly",
                    metric=metric, value=value, unit=unit, source_row=row_number, raw=evidence,
                    quality_level="moderate", methodology=methodology +
                    " Workbook revisions are preserved as separate observations; release dates are unavailable in this historical table.",
                )
                yield row
        finally:
            workbook.close()
